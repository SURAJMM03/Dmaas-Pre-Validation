"""
DMaaS — Data Migration 
Prevalidation Engine  |  Powered by Groq AI (Free)
"""

import streamlit as st
import pandas as pd
import re, os, math, json, httpx
from io import BytesIO
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from groq import Groq as _GroqClient

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DMaaS · Prevalidation",
    layout="wide",
    page_icon="⚡",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────
# DESIGN SYSTEM
# ─────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;600&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; color: #e2e8f0; }
.main { background: #080c14; }
.block-container { padding: 2rem 2.5rem 4rem; max-width: 1400px; }

section[data-testid="stSidebar"] { background: #0d1117; border-right: 1px solid #1e2a3a; }
section[data-testid="stSidebar"] .block-container { padding: 1.5rem 1.2rem; }

/* HERO */
.hero-wrap { display:flex; align-items:flex-end; justify-content:space-between;
    padding:2rem 0 1.6rem; border-bottom:1px solid #1e2a3a; margin-bottom:2rem; }
.hero-brand { font-family:'JetBrains Mono',monospace; font-size:0.72rem; font-weight:600;
    letter-spacing:3px; text-transform:uppercase; color:#3b82f6; margin-bottom:0.5rem; }
.hero-title { font-size:2.2rem; font-weight:700; color:#f1f5f9;
    letter-spacing:-0.5px; line-height:1.15; margin:0; }
.hero-sub { font-size:0.9rem; color:#64748b; margin-top:0.4rem; }
.hero-badge { display:inline-flex; align-items:center; gap:6px; background:#0f2027;
    border:1px solid #1e3a5f; border-radius:20px; padding:6px 14px;
    font-size:0.75rem; font-weight:500; color:#60a5fa; font-family:'JetBrains Mono',monospace; }
.hero-dot { width:7px; height:7px; border-radius:50%; background:#22c55e;
    box-shadow:0 0 8px #22c55e; animation:pulse 2s infinite; }
@keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.4} }

/* SECTION LABEL */
.sec-label { font-family:'JetBrains Mono',monospace; font-size:0.65rem;
    text-transform:uppercase; letter-spacing:3px; color:#3b82f6;
    margin-bottom:1rem; padding-bottom:6px; border-bottom:1px solid #1e2a3a;
    display:flex; align-items:center; gap:8px; }
.sec-label::before { content:''; display:block; width:3px; height:12px;
    background:#3b82f6; border-radius:2px; }

/* KPI CARDS */
.kpi-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin-bottom:1.5rem; }
.kpi { background:#0d1117; border:1px solid #1e2a3a; border-radius:12px;
    padding:1.2rem 1.4rem; position:relative; overflow:hidden; }
.kpi::before { content:''; position:absolute; top:0; left:0; right:0; height:2px; }
.kpi.red::before   { background:linear-gradient(90deg,#ef4444,#f97316); }
.kpi.amber::before { background:linear-gradient(90deg,#f59e0b,#eab308); }
.kpi.green::before { background:linear-gradient(90deg,#22c55e,#10b981); }
.kpi-num { font-family:'JetBrains Mono',monospace; font-size:2.4rem;
    font-weight:700; line-height:1; margin-bottom:0.3rem; }
.kpi.red   .kpi-num { color:#f87171; }
.kpi.amber .kpi-num { color:#fbbf24; }
.kpi.green .kpi-num { color:#4ade80; }
.kpi-lbl { font-size:0.7rem; text-transform:uppercase; letter-spacing:1.5px;
    color:#475569; font-weight:500; }
.kpi-sub { font-size:0.75rem; color:#374151; margin-top:0.25rem; }

/* SHEET TABLE */
.sheet-table { width:100%; border-collapse:collapse; margin-top:0.5rem; }
.sheet-table th { background:#0d1117; color:#64748b; font-size:0.7rem; font-weight:600;
    text-transform:uppercase; letter-spacing:1px; padding:10px 14px;
    border-bottom:1px solid #1e2a3a; text-align:left; }
.sheet-table td { padding:10px 14px; border-bottom:1px solid #111827;
    font-size:0.82rem; color:#cbd5e1; }
.sheet-table tr:last-child td { border-bottom:none; }
.badge { display:inline-block; padding:2px 9px; border-radius:20px;
    font-size:0.68rem; font-weight:600; letter-spacing:.5px; }
.badge.pass { background:#052e16; color:#4ade80; border:1px solid #166534; }
.badge.fail { background:#2d0a0a; color:#f87171; border:1px solid #7f1d1d; }

/* RISK ROWS */
.risk-row { display:flex; align-items:center; gap:10px; background:#0a0f1a;
    border:1px solid #1e2a3a; border-radius:8px; padding:8px 14px; margin-bottom:6px; }
.risk-bar-wrap { flex:1; background:#111827; border-radius:4px; height:6px; }
.risk-bar { height:6px; border-radius:4px; }
.risk-bar.high { background:linear-gradient(90deg,#ef4444,#f97316); }
.risk-bar.med  { background:linear-gradient(90deg,#f59e0b,#eab308); }
.risk-meta { font-size:0.75rem; font-family:'JetBrains Mono',monospace;
    color:#64748b; width:48px; text-align:right; }

/* AI CARDS */
.ai-card { background:#0d1117; border:1px solid #1e2a3a; border-radius:14px;
    padding:1.4rem 1.6rem; margin-bottom:1rem; }
.ai-card-title { font-size:0.78rem; font-weight:600; text-transform:uppercase;
    letter-spacing:1.5px; color:#60a5fa; margin-bottom:0.8rem;
    display:flex; align-items:center; gap:8px; }
.ai-card-body { font-size:0.875rem; color:#94a3b8; line-height:1.8; }
.ai-card-body strong { color:#e2e8f0; }

/* CHECKLIST */
.chk-item { display:flex; align-items:flex-start; gap:12px; background:#0a0f1a;
    border:1px solid #1e2a3a; border-radius:10px; padding:12px 16px; margin-bottom:8px; }
.chk-item.pass { border-left:3px solid #22c55e; }
.chk-item.fail { border-left:3px solid #ef4444; }
.chk-item.warn { border-left:3px solid #f59e0b; }
.chk-status { font-size:1rem; flex-shrink:0; margin-top:1px; }
.chk-title  { font-size:0.82rem; font-weight:600; color:#e2e8f0; }
.chk-reason { font-size:0.75rem; color:#64748b; margin-top:2px; }

/* FIX GUIDE CARDS */
.fix-card { background:#0a0f1a; border:1px solid #1e2a3a; border-radius:10px;
    padding:14px 16px; margin-bottom:10px; }
.fix-header { display:flex; align-items:center; gap:10px; margin-bottom:8px; }
.fix-badge { padding:3px 10px; border-radius:6px; font-size:0.68rem;
    font-weight:700; letter-spacing:.5px; }
.fix-badge.mand { background:#2d0a0a; color:#f87171; border:1px solid #7f1d1d; }
.fix-badge.date { background:#1c1409; color:#fbbf24; border:1px solid #92400e; }
.fix-badge.num  { background:#0a1f2a; color:#38bdf8; border:1px solid #1e3a5f; }
.fix-badge.len  { background:#1a0f2e; color:#a78bfa; border:1px solid #4c1d95; }
.fix-field { font-size:0.85rem; font-weight:600; color:#e2e8f0; }
.fix-row { font-size:0.78rem; color:#64748b; margin-top:4px; line-height:1.65; }
.fix-row strong { color:#94a3b8; }

/* VERDICT BANNERS */
.verdict-pass { background:#052e16; border:1px solid #166534; border-radius:10px;
    padding:14px 18px; color:#4ade80; font-weight:600; font-size:0.9rem; margin:0.5rem 0 1rem; }
.verdict-warn { background:#1c1409; border:1px solid #92400e; border-radius:10px;
    padding:14px 18px; color:#fbbf24; font-weight:600; font-size:0.9rem; margin:0.5rem 0 1rem; }
.verdict-fail { background:#2d0a0a; border:1px solid #7f1d1d; border-radius:10px;
    padding:14px 18px; color:#f87171; font-weight:600; font-size:0.9rem; margin:0.5rem 0 1rem; }

/* MISC */
div[data-testid="stFileUploader"] { background:#0d1117; border:1.5px dashed #1e3a5f;
    border-radius:12px; padding:0.5rem; }
.stButton > button { background:linear-gradient(135deg,#1d4ed8,#1e40af); color:white;
    border:none; border-radius:10px; padding:0.65rem 2rem;
    font-weight:600; font-size:0.9rem; width:100%; transition:all .2s; }
.stButton > button:hover { background:linear-gradient(135deg,#2563eb,#1d4ed8);
    transform:translateY(-1px); box-shadow:0 4px 20px rgba(59,130,246,.3); }
.stDownloadButton > button { background:#0a1f0a; color:#4ade80; border:1px solid #166534;
    border-radius:8px; font-size:0.8rem; font-weight:500; width:100%; }
.log-box { background:#030712; border:1px solid #111827; border-radius:10px; padding:1rem;
    font-family:'JetBrains Mono',monospace; font-size:0.72rem; color:#475569;
    max-height:240px; overflow-y:auto; line-height:1.6; }
.stTabs [data-baseweb="tab-list"] { background:#0d1117; border-bottom:1px solid #1e2a3a; }
.stTabs [data-baseweb="tab"] { background:transparent; border:none; color:#64748b;
    font-size:0.8rem; font-weight:500; padding:8px 16px; }
.stTabs [aria-selected="true"] { background:#111827 !important; color:#e2e8f0 !important;
    border-bottom:2px solid #3b82f6 !important; }
.stTabs [data-baseweb="tab-panel"] { padding-top:1.2rem; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# GROQ CONFIG  ← paste your gsk_... key here
# ─────────────────────────────────────────────────────────────
_GROQ_API_KEY        = "API_KEY_HERE"
_GROQ_PRIMARY_MODEL  = "llama-3.3-70b-versatile"
_GROQ_FALLBACK_MODEL = "llama-3.1-8b-instant"
_GROQ_TEMPERATURE    = 0.15
_GROQ_TIMEOUT        = 60
_groq_cache: dict    = {}
_SYSTEM = ("You are a senior SAP Data Migration consultant. "
           "Be precise, concise, and enterprise-ready. No filler words.")


def _resolve_key() -> str:
    if _GROQ_API_KEY and _GROQ_API_KEY != "PASTE_YOUR_GROQ_KEY_HERE":
        return _GROQ_API_KEY
    try:
        k = st.secrets.get("GROQ_API_KEY", "")
        if k: return k
    except Exception:
        pass
    return os.environ.get("GROQ_API_KEY", st.session_state.get("_sidebar_key", ""))


def _client():
    key = _resolve_key()
    if not key: return None
    if key not in _groq_cache:
        _groq_cache.clear()
        _groq_cache[key] = _GroqClient(
            api_key=key, timeout=_GROQ_TIMEOUT,
            http_client=httpx.Client(verify=False))
    return _groq_cache[key]


def call_llm(prompt: str, max_tokens: int = 500,
             json_mode: bool = False, model: str = _GROQ_PRIMARY_MODEL) -> str:
    c = _client()
    if c is None: return "⚠️ Groq API key not configured."
    kw = dict(model=model, max_tokens=max_tokens, temperature=_GROQ_TEMPERATURE,
              messages=[{"role": "system", "content": _SYSTEM},
                        {"role": "user",   "content": prompt}])
    if json_mode: kw["response_format"] = {"type": "json_object"}
    try:
        r = c.chat.completions.create(**kw)
        return (r.choices[0].message.content or "").strip()
    except Exception as ex:
        err = str(ex)
        if ("429" in err or "rate_limit" in err.lower()) and model == _GROQ_PRIMARY_MODEL:
            try:
                kw["model"] = _GROQ_FALLBACK_MODEL
                r2 = c.chat.completions.create(**kw)
                return (r2.choices[0].message.content or "").strip()
            except Exception:
                return "⚠️ Rate limit hit — try again in a moment."
        if "401" in err or "invalid_api_key" in err.lower():
            return "⚠️ Invalid API key — check _GROQ_API_KEY in the code."
        if "connect" in err.lower() or "timeout" in err.lower():
            return "⚠️ Connection error — check your internet."
        return f"⚠️ AI error: {err[:120]}"


# ─────────────────────────────────────────────────────────────
# VALIDATION HELPERS
# ─────────────────────────────────────────────────────────────
def strip_star(h: str):
    h = str(h).split('\n')[0].strip()
    return h.replace('*', '').strip(), '*' in h

def norm(t: str) -> str:
    return re.sub(r'\s+', ' ', str(t).strip()).lower()

def fuzzy_score(a: str, b: str) -> float:
    an, bn = norm(a), norm(b)
    s = SequenceMatcher(None, an, bn).ratio()
    if an in bn or bn in an: s = max(s, 0.80)
    return s

def fuzzy(q: str, candidates: list, threshold: float = 0.72):
    qn = norm(q); best, bs = None, 0.0
    for c in candidates:
        cn = norm(str(c))
        if qn == cn: return c
        s = SequenceMatcher(None, qn, cn).ratio()
        if qn in cn or cn in qn: s = max(s, 0.80)
        if s > bs: bs, best = s, c
    return best if bs >= threshold else None

def best_col(hdrs: list, aliases: list, th: float = 0.65):
    bi, bs = None, 0.0
    for i, h in enumerate(hdrs):
        if not h: continue
        for a in aliases:
            s = fuzzy_score(h, a)
            if s > bs: bs, bi = s, i
    return bi if bs >= th else None

def is_blank(v) -> bool:
    if v is None: return True
    try:
        if v != v: return True
    except Exception: pass
    try:
        if isinstance(v, float) and math.isnan(v): return True
    except Exception: pass
    return str(v).strip() in ('', 'nan', 'none', 'NaN', 'None')

def check_date(v) -> bool:
    if isinstance(v, datetime): return True
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none', ''): return True
    try: datetime.strptime(s, '%d.%m.%Y'); return True
    except ValueError: return False

def check_time(v) -> bool:
    if isinstance(v, datetime): return True
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none', ''): return True
    if re.match(r'^\d{6}$', s):
        hh, mm, ss = int(s[0:2]), int(s[2:4]), int(s[4:6])
        return 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59
    m = re.match(r'^(\d{1,2}):(\d{2}):(\d{2})$', s)
    if m:
        hh, mm, ss = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return 0 <= hh <= 23 and 0 <= mm <= 59 and 0 <= ss <= 59
    return False

def check_number(v, max_len, dec_places) -> tuple:
    if is_blank(v): return True, ''
    s = str(v).strip()
    try: num = float(s)
    except: return False, f"'{s}' is not a valid number"
    whole = dec_places is None or dec_places == 0 or str(dec_places).strip() in ('', '0')
    if whole:
        if not float(s).is_integer(): return False, f"'{s}' must be a whole number"
        d = len(str(abs(int(num))))
        if max_len and d > max_len: return False, f"'{s}' has {d} digits (max {max_len})"
    else:
        dp = int(dec_places)
        if '.' in s:
            adp = len(s.split('.')[1])
            if adp > dp: return False, f"'{s}' has {adp} decimal places (max {dp})"
        ip = str(abs(int(num)))
        if max_len and len(ip) > max_len:
            return False, f"'{s}' integer part too long (max {max_len})"
    return True, ''


# ─────────────────────────────────────────────────────────────
# FILLS
# ─────────────────────────────────────────────────────────────
FILLS = {
    'RED':    PatternFill('solid', start_color='DC2626', end_color='DC2626'),
    'YELLOW': PatternFill('solid', start_color='D97706', end_color='D97706'),
    'TEAL':   PatternFill('solid', start_color='0891B2', end_color='0891B2'),
    'ORANGE': PatternFill('solid', start_color='EA580C', end_color='EA580C'),
}
HDR_FILL = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
HDR_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=10)


# ─────────────────────────────────────────────────────────────
# LOAD RULES
# ─────────────────────────────────────────────────────────────
def load_rules(rules_bytes: bytes, log: list) -> dict:
    wb = load_workbook(BytesIO(rules_bytes), data_only=True)
    all_rules = defaultdict(dict)
    COL_ALIASES = {
        'sheet_name':    ['Sheet Name', 'Sheet', 'Tab Name', 'Tab'],
        'field_desc':    ['Field Description', 'Field Desc', 'Field Name', 'Description'],
        'importance':    ['Importance', 'Mandatory', 'Required', 'Compulsory'],
        'type':          ['Type', 'Data Type', 'Field Type'],
        'length':        ['Length', 'Max Length', 'Field Length', 'Size'],
        'decimal':       ['Decimal', 'Decimals', 'Decimal Places'],
        'sap_field':     ['SAP Field', 'SAP Fieldname'],
        'sap_structure': ['SAP Structure', 'Structure'],
    }
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        hdr_row, found = 8, False
        for ri in range(1, min(20, ws.max_row + 1)):
            for ci in range(1, ws.max_column + 1):
                if 'field description' in norm(str(ws.cell(ri, ci).value or '')):
                    hdr_row, found = ri, True; break
            if found: break
        log.append(f"  Rules sheet '{ws_name}': header row {hdr_row}")
        raw_hdrs = [str(ws.cell(hdr_row, c).value or '').strip()
                    for c in range(1, ws.max_column + 1)]
        col_idx: dict = {}
        for key, aliases in COL_ALIASES.items():
            idx = best_col(raw_hdrs, aliases)
            if idx is not None: col_idx[key] = idx

        def make_gcell(snap):
            def gcell(row, key):
                idx = snap.get(key)
                if idx is None or idx >= len(row): return ''
                v = row[idx]
                return str(v).strip() if v is not None else ''
            return gcell

        gcell = make_gcell(col_idx)
        count = 0
        for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
            fd = gcell(row, 'field_desc')
            sn = gcell(row, 'sheet_name')
            if not fd or fd.lower() in ('none', 'nan', ''): continue
            lr = gcell(row, 'length'); dr = gcell(row, 'decimal')
            try:    ml = int(float(lr)) if lr else None
            except: ml = None
            try:    dp = int(float(dr)) if dr and dr not in ('', 'none', 'nan') else None
            except: dp = None
            tlc = gcell(row, 'type').strip().lower()
            rule = {
                'field_desc': fd, 'sheet_name': sn,
                'is_mandatory': gcell(row, 'importance').lower().strip() == 'mandatory for sheet',
                'type': gcell(row, 'type').strip(),
                'is_date':   tlc == 'date',
                'is_time':   tlc in ('time', 'tims'),
                'is_number': tlc in ('number', 'num', 'integer', 'int', 'float', 'decimal',
                                     'amount', 'quantity', 'curr', 'quan'),
                'is_text':   tlc in ('text', 'char', 'string'),
                'max_length': ml, 'decimal_places': dp,
                'sap_field': gcell(row, 'sap_field'),
            }
            sk = norm(sn) if sn else '__global__'
            all_rules[sk][norm(fd)] = rule
            count += 1
        log.append(f"    {count} rules loaded")
    return dict(all_rules)


# ─────────────────────────────────────────────────────────────
# HEADER DETECTION
# ─────────────────────────────────────────────────────────────
def detect_header_row(ws, rule_fields: list, fixed: int = 8) -> int:
    candidates = ([fixed]
                  + list(range(max(1, fixed - 3), fixed))
                  + list(range(fixed + 1, min(fixed + 6, ws.max_row + 1))))
    best_row, best_score = fixed, 0
    for ri in candidates:
        vals = [str(ws.cell(ri, ci).value or '').strip()
                for ci in range(1, ws.max_column + 1)]
        if not any(vals): continue
        score = sum(1 for h in vals if h and fuzzy(norm(h), rule_fields, 0.60))
        if score > best_score: best_score, best_row = score, ri
    return best_row


# ─────────────────────────────────────────────────────────────
# VALIDATE SHEET
# ─────────────────────────────────────────────────────────────
def validate_sheet(ws, sheet_name: str, all_rules: dict, log: list) -> tuple:
    errors = []; skey = norm(sheet_name)
    rules = all_rules.get(skey, {})
    if not rules:
        for rk in all_rules:
            if rk != '__global__' and SequenceMatcher(None, skey, rk).ratio() > 0.75:
                rules = all_rules[rk]; log.append(f"  Fuzzy matched '{rk}'"); break
    merged = {**all_rules.get('__global__', {}), **rules}
    if not merged: log.append(f"  No rules — skipping '{sheet_name}'"); return errors, 0
    hdr = detect_header_row(ws, list(merged.keys()))
    log.append(f"  Header row {hdr}")
    has_data = any(ws.cell(ri, ci).value is not None
                   for ri in range(hdr + 1, ws.max_row + 1)
                   for ci in range(1, ws.max_column + 1))
    if not has_data: log.append(f"  No data — skipping"); return errors, 0
    col_info: dict = {}
    for ci in range(1, ws.max_column + 1):
        raw = ws.cell(hdr, ci).value
        if raw is not None:
            clean, star = strip_star(str(raw))
            if clean: col_info[ci] = {'clean': clean, 'star': star}
    fkeys = list(merged.keys())
    c2r: dict = {}
    for ci, info in col_info.items():
        mk = fuzzy(norm(info['clean']), fkeys, 0.72)
        if mk:
            rule = merged[mk].copy()
            if info['star']: rule['is_mandatory'] = True
            c2r[ci] = (info, rule)
    log.append(f"  {len(c2r)}/{len(col_info)} columns matched")
    id_col = None
    for ci, (info, _) in c2r.items():
        if fuzzy(norm(info['clean']),
                 ['product number', 'material number', 'order number', 'matnr', 'id', 'key'], 0.65):
            id_col = ci; break
    rows_checked = 0
    for ri in range(hdr + 1, ws.max_row + 1):
        if all(is_blank(ws.cell(ri, ci).value) for ci in col_info): continue
        rows_checked += 1
        id_val = ''
        if id_col:
            iv = ws.cell(ri, id_col).value
            id_val = str(iv).strip() if not is_blank(iv) else ''
        for ci, (info, rule) in c2r.items():
            cell_val = ws.cell(ri, ci).value
            empty    = is_blank(cell_val)
            val_str  = '' if empty else str(cell_val).strip()
            def make_err(sh, rw, prod, fld, val, col):
                def err(etype, msg, hl):
                    return {'Sheet Name': sh, 'Row Number': rw, 'Product Number': prod,
                            'Field Name': fld, 'Field Value': val,
                            'Error Type': etype, 'Error Message': msg,
                            '_col': col, '_hl': hl}
                return err
            err = make_err(sheet_name, ri, id_val, info['clean'], val_str, ci)
            if rule['is_mandatory'] and empty:
                errors.append(err('Mandatory Field Missing', 'Mandatory field missing', 'RED'))
                continue
            if empty: continue
            if rule['is_text'] and rule['max_length'] and not rule['is_date']:
                if len(val_str) > rule['max_length']:
                    errors.append(err('Length Exceeded',
                        f"Length {len(val_str)} exceeds max {rule['max_length']}", 'YELLOW'))
            if rule['is_date'] and not check_date(cell_val):
                errors.append(err('Invalid Date Format',
                    f"'{val_str}' — expected dd.mm.yyyy", 'YELLOW'))
            if rule['is_time'] and not check_time(cell_val):
                errors.append(err('Invalid Time Format',
                    f"'{val_str}' — expected HHMMSS or HH:MM:SS", 'TEAL'))
            if rule['is_number']:
                ok, detail = check_number(cell_val, rule['max_length'], rule['decimal_places'])
                if not ok: errors.append(err('Invalid Number', detail, 'ORANGE'))
    log.append(f"  {rows_checked} rows checked, {len(errors)} error(s)")
    return errors, rows_checked


# ─────────────────────────────────────────────────────────────
# ROW RISK SCORING  (no LLM — deterministic)
# ─────────────────────────────────────────────────────────────
def compute_row_risk(all_errors: list) -> dict:
    W = {'Mandatory Field Missing': 40, 'Invalid Date Format': 20,
         'Invalid Number': 20, 'Length Exceeded': 10, 'Invalid Time Format': 10}
    row_errs = defaultdict(list)
    for e in all_errors: row_errs[(e['Sheet Name'], e['Row Number'])].append(e)
    return {k: min(100, sum(W.get(e['Error Type'], 5) for e in v))
            for k, v in row_errs.items()}


# ─────────────────────────────────────────────────────────────
# BUILD HIGHLIGHTED XLSX
# ─────────────────────────────────────────────────────────────
def build_highlighted(product_bytes: bytes, errors: list, all_rules: dict) -> bytes:
    wb = load_workbook(BytesIO(product_bytes))
    pri = {'RED': 4, 'TEAL': 3, 'ORANGE': 2, 'YELLOW': 1}
    hmap: dict = {}
    for e in errors:
        k = (e['Sheet Name'], e['Row Number'], e['_col'])
        if pri.get(e['_hl'], 0) > pri.get(hmap.get(k, ''), 0): hmap[k] = e['_hl']
    for sn in wb.sheetnames:
        ws = wb[sn]; sk = norm(sn)
        rules = all_rules.get(sk, {})
        if not rules:
            for rk in all_rules:
                if rk != '__global__' and SequenceMatcher(None, sk, rk).ratio() > 0.75:
                    rules = all_rules[rk]; break
        merged  = {**all_rules.get('__global__', {}), **rules}
        hdr_row = detect_header_row(ws, list(merged.keys())) if merged else 8
        for ci in range(1, ws.max_column + 1):
            if ws.cell(hdr_row, ci).value is not None:
                ws.cell(hdr_row, ci).fill = HDR_FILL
                ws.cell(hdr_row, ci).font = HDR_FONT
        for (s, r, c), colour in hmap.items():
            if s == sn: ws.cell(r, c).fill = FILLS[colour]
    leg = wb.create_sheet('_Legend')
    leg['A1'].value = 'Highlight Legend'
    leg['A1'].font  = Font(bold=True, name='Calibri', size=12, color='FFFFFF')
    leg['A1'].fill  = HDR_FILL
    for i, (label, col) in enumerate(
            [('Mandatory field empty', 'RED'), ('Length / Date error', 'YELLOW'),
             ('Time format error', 'TEAL'), ('Number / Decimal error', 'ORANGE')], 2):
        leg.cell(i, 1).value = label
        leg.cell(i, 1).fill  = FILLS[col]
        leg.cell(i, 1).font  = Font(color='FFFFFF', name='Calibri', size=10)
    leg.column_dimensions['A'].width = 30
    out = BytesIO(); wb.save(out); return out.getvalue()


# ─────────────────────────────────────────────────────────────
# BUILD ERROR REPORT XLSX
# ─────────────────────────────────────────────────────────────
def build_error_report(errors: list, filename: str,
                       sheet_row_counts: dict = None,
                       row_risk: dict = None) -> bytes:
    from openpyxl import Workbook as WB
    HF   = PatternFill('solid', start_color='1E3A5F', end_color='1E3A5F')
    HFT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
    thin = Side(style='thin', color='1E2A3A')
    BDR  = Border(left=thin, right=thin, top=thin, bottom=thin)
    NF   = Font(name='Calibri', size=10)
    TF   = {
        'Mandatory Field Missing': PatternFill('solid', start_color='FFE8E8', end_color='FFE8E8'),
        'Length Exceeded':         PatternFill('solid', start_color='FFFBE6', end_color='FFFBE6'),
        'Invalid Date Format':     PatternFill('solid', start_color='FFFBE6', end_color='FFFBE6'),
        'Invalid Time Format':     PatternFill('solid', start_color='E0F7F5', end_color='E0F7F5'),
        'Invalid Number':          PatternFill('solid', start_color='FFF0E6', end_color='FFF0E6'),
    }
    GRN = PatternFill('solid', start_color='F0FDF4', end_color='F0FDF4')
    RED = PatternFill('solid', start_color='FEF2F2', end_color='FEF2F2')
    wb = WB(); ws_s = wb.active; ws_s.title = 'Summary'
    ws_s.merge_cells('A1:J1')
    ws_s['A1'].value = f'Validation Report  ·  {filename}  ·  {datetime.now().strftime("%d.%m.%Y %H:%M")}'
    ws_s['A1'].font  = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
    ws_s['A1'].fill  = HF
    ws_s['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_s.row_dimensions[1].height = 28
    counts = Counter(e['Error Type'] for e in errors)
    for ci, (lbl, val, col) in enumerate(
        [('Total Errors', len(errors), 'CC0000'),
         ('Mandatory', counts.get('Mandatory Field Missing', 0), 'CC0000'),
         ('Length',    counts.get('Length Exceeded', 0), 'CC6600'),
         ('Date',      counts.get('Invalid Date Format', 0), 'CC6600'),
         ('Time',      counts.get('Invalid Time Format', 0), '0891B2'),
         ('Number',    counts.get('Invalid Number', 0), 'EA580C')], 1):
        ws_s.cell(3, ci).value = lbl
        ws_s.cell(3, ci).font  = Font(bold=True, size=9, color='555555', name='Calibri')
        ws_s.cell(4, ci).value = val
        ws_s.cell(4, ci).font  = Font(bold=True, size=20, color=col, name='Calibri')
        ws_s.column_dimensions[get_column_letter(ci)].width = 18
    ws_s.cell(6, 1).value = 'Error Summary by Sheet'
    ws_s.cell(6, 1).font  = Font(bold=True, name='Calibri')
    for ci, h in enumerate(['Sheet', 'Total Rows', 'Clean Rows', 'Error Rows', 'Total Errors',
                             'Mandatory', 'Length', 'Date', 'Time', 'Number'], 1):
        c = ws_s.cell(7, ci, value=h)
        c.fill, c.font, c.border = HF, HFT, BDR
        ws_s.column_dimensions[get_column_letter(ci)].width = 16
    src = sheet_row_counts or {}
    ser = defaultdict(set); ss = defaultdict(lambda: defaultdict(int))
    for e in errors:
        s = e['Sheet Name']; ser[s].add(e['Row Number'])
        ss[s]['total'] += 1; ss[s][e['Error Type']] += 1
    for ri, sn in enumerate(list(src.keys()) or list(ss.keys()), 8):
        tr = src.get(sn, ''); er = len(ser[sn])
        cr = (tr - er) if isinstance(tr, int) else ''
        sc = ss.get(sn, {}); rf = GRN if er == 0 else RED
        for ci, v in enumerate(
            [sn, tr, cr, er, sc.get('total', 0),
             sc.get('Mandatory Field Missing', 0), sc.get('Length Exceeded', 0),
             sc.get('Invalid Date Format', 0), sc.get('Invalid Time Format', 0),
             sc.get('Invalid Number', 0)], 1):
            c = ws_s.cell(ri, ci, value=v); c.font, c.border, c.fill = NF, BDR, rf
    ws_d = wb.create_sheet('Error Details')
    OCOLS = ['Sheet Name', 'Row Number', 'Product Number', 'Field Name',
             'Field Value', 'Error Type', 'Error Message', 'Risk Score']
    WIDS  = [20, 10, 22, 35, 28, 24, 60, 12]
    for ci, (h, w) in enumerate(zip(OCOLS, WIDS), 1):
        c = ws_d.cell(1, ci, value=h)
        c.fill, c.font, c.border = HF, HFT, BDR
        c.alignment = Alignment(horizontal='center')
        ws_d.column_dimensions[get_column_letter(ci)].width = w
    ws_d.freeze_panes = 'A2'
    rr = row_risk or {}
    for ri, e in enumerate(errors, 2):
        fill  = TF.get(e['Error Type'], PatternFill())
        score = rr.get((e['Sheet Name'], e['Row Number']), '')
        for ci, key in enumerate(OCOLS[:-1], 1):
            c = ws_d.cell(ri, ci, value=e.get(key, ''))
            c.fill, c.font, c.border = fill, NF, BDR
        c8 = ws_d.cell(ri, 8, value=score); c8.fill, c8.font, c8.border = fill, NF, BDR
    ws_d.auto_filter.ref = f"A1:{get_column_letter(len(OCOLS))}1"
    out = BytesIO(); wb.save(out); return out.getvalue()


# ═══════════════════════════════════════════════════════════════
# AI FEATURES  — 5 focused, high-value features
# ═══════════════════════════════════════════════════════════════

def ai_health_brief(all_errors, sheet_row_counts, filename) -> str:
    """5-line executive brief: health, risk level, go/no-go, biggest issue, today's action."""
    total = sum(sheet_row_counts.values())
    erws  = len(set((e['Sheet Name'], e['Row Number']) for e in all_errors))
    cnt   = Counter(e['Error Type'] for e in all_errors)
    rate  = round(erws / max(total, 1) * 100, 1)
    sheets = ' | '.join(
        f"{sn}: {sum(1 for e in all_errors if e['Sheet Name']==sn)} err/{rc} rows"
        for sn, rc in sheet_row_counts.items())
    prompt = (
        f"SAP migration validation for '{filename}':\n"
        f"Rows: {total} total | {total-erws} clean | {erws} error rows ({rate}%)\n"
        f"Mandatory={cnt.get('Mandatory Field Missing',0)} "
        f"Date={cnt.get('Invalid Date Format',0)} "
        f"Number={cnt.get('Invalid Number',0)} "
        f"Length={cnt.get('Length Exceeded',0)} "
        f"Time={cnt.get('Invalid Time Format',0)}\n"
        f"Sheets: {sheets}\n\n"
        f"Write exactly 5 plain lines. No bullets. No labels. No markdown:\n"
        f"1. Overall data health status (1 sentence)\n"
        f"2. RISK LEVEL: CRITICAL/HIGH/MEDIUM/LOW — reason (8 words)\n"
        f"3. MIGRATION: GO/CONDITIONAL GO/NO-GO — reason (8 words)\n"
        f"4. The single most critical problem to fix (be specific)\n"
        f"5. The one action the team must do today"
    )
    return call_llm(prompt, 280)


def ai_fix_guide(all_errors) -> dict:
    """Top 8 error patterns: why it happens, how to fix, how to verify."""
    top = Counter((e['Error Type'], e['Field Name']) for e in all_errors).most_common(8)
    if not top: return {}
    lines = '\n'.join(f"{et} | {fn} | {cnt}" for (et, fn), cnt in top)
    prompt = (
        f"SAP migration errors:\n{lines}\n\n"
        f"Return JSON. For each 'ErrorType|FieldName' key return an object:\n"
        f"  'why': root cause in SAP migrations (max 12 words)\n"
        f"  'fix': corrective action (max 20 words)\n"
        f"  'verify': how to confirm fixed (max 12 words)\n"
        f"Keys must exactly match 'ErrorType|FieldName' above. Return ONLY valid JSON."
    )
    raw = call_llm(prompt, 800, json_mode=True)
    result: dict = {}
    try:
        parsed = json.loads(re.sub(r'```(?:json)?|```', '', raw).strip())
        for (et, fn), _ in top:
            k = f"{et}|{fn}"
            if k in parsed: result[(et, fn)] = parsed[k]
    except Exception:
        pass
    return result


def ai_readiness_checklist(all_errors, sheet_row_counts, filename) -> list:
    """8-point PASS/WARN/FAIL go/no-go checklist with specific numbers."""
    total = sum(sheet_row_counts.values())
    erws  = len(set((e['Sheet Name'], e['Row Number']) for e in all_errors))
    rate  = round(erws / max(total, 1) * 100, 1)
    cnt   = Counter(e['Error Type'] for e in all_errors)
    prompt = (
        f"SAP migration go/no-go for '{filename}':\n"
        f"Rows:{total} | Error rows:{erws} | Rate:{rate}%\n"
        f"Mandatory:{cnt.get('Mandatory Field Missing',0)} "
        f"Date:{cnt.get('Invalid Date Format',0)} "
        f"Number:{cnt.get('Invalid Number',0)} "
        f"Length:{cnt.get('Length Exceeded',0)}\n\n"
        f"Return a JSON array of exactly 8 checklist items covering:\n"
        f"mandatory_completeness, date_format, number_format, field_length,\n"
        f"overall_error_rate, critical_error_absence, data_volume, migration_risk\n"
        f"Rules: FAIL if mandatory>0 or rate>15%. WARN if rate 5-15% or non-critical errors. PASS otherwise.\n"
        f"Each: {{\"item\":\"plain label\",\"status\":\"PASS|WARN|FAIL\","
        f"\"reason\":\"one sentence with actual numbers\"}}\n"
        f"Return ONLY a valid JSON array."
    )
    raw = call_llm(prompt, 700, json_mode=True)
    try:
        cleaned = re.sub(r'```(?:json)?|```', '', raw).strip()
        result  = json.loads(cleaned)
        if isinstance(result, list): return result
        for v in result.values():
            if isinstance(v, list): return v
    except Exception:
        pass
    return []


def ai_root_cause(all_errors, sheet_row_counts) -> str:
    """Analyses actual bad values to diagnose root causes and recommend bulk fixes."""
    if not all_errors: return "No errors to analyse."
    samples = []
    for e in all_errors:
        v = e.get('Field Value', '')
        if v and v not in ('', 'nan', 'None'):
            samples.append(f"{e['Field Name']} | {e['Error Type']} | '{v}'")
        if len(samples) >= 35: break
    sheet_rates = {sn: f"{sum(1 for e in all_errors if e['Sheet Name']==sn)}/{sheet_row_counts.get(sn,1)}"
                   for sn in Counter(e['Sheet Name'] for e in all_errors)}
    top_fields  = Counter(e['Field Name'] for e in all_errors).most_common(6)
    prompt = (
        f"Diagnose root causes of these SAP migration errors.\n\n"
        f"SAMPLE BAD VALUES:\n" + '\n'.join(samples) +
        f"\n\nError rate by sheet: {sheet_rates}"
        f"\nTop error fields: {top_fields}\n\n"
        f"Write under exactly these 3 plain-text headers "
        f"(no markdown, separate with blank lines):\n\n"
        f"WHAT IS WRONG\n"
        f"Top 3 error patterns from the actual values above. Quote specific values.\n\n"
        f"WHY IT IS HAPPENING\n"
        f"Most likely root causes — source system, regional formats, Excel, data entry.\n\n"
        f"HOW TO FIX IT AT SCALE\n"
        f"Bulk transformations the team can apply. Name specific tools or steps."
    )
    return call_llm(prompt, 500)


def ai_chat(question, all_errors, sheet_row_counts, filename) -> str:
    """Data analyst chat — answers questions about validation results."""
    if not question.strip(): return ''
    cnt    = Counter(e['Error Type'] for e in all_errors)
    by_sht = Counter(e['Sheet Name']  for e in all_errors)
    by_fld = Counter(e['Field Name']  for e in all_errors).most_common(10)
    sample = [f"Row {e['Row Number']} | {e['Sheet Name']} | {e['Field Name']} | "
              f"{e['Error Type']} | '{e.get('Field Value', '')}'"
              for e in all_errors[:25]]
    prompt = (
        f"You are a data analyst. Answer questions about SAP migration validation.\n"
        f"File: {filename} | Rows: {sum(sheet_row_counts.values())} | "
        f"Errors: {len(all_errors)}\n"
        f"By type: {dict(cnt)} | By sheet: {dict(by_sht)}\n"
        f"Top fields: {by_fld} | Sheet sizes: {sheet_row_counts}\n"
        f"Sample errors:\n" + '\n'.join(sample) +
        f"\n\nQUESTION: {question}\n\n"
        f"Answer directly. Show arithmetic if calculating. Max 10 lines."
    )
    return call_llm(prompt, 400)


# ─────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────
for _k in ('all_errors', 'log', 'hl_bytes', 'err_bytes', 'csv_bytes', 'base_name',
           'log_txt_bytes', 'log_filename', 'sheet_row_counts', 'row_risk',
           'ai_brief', 'ai_fix', 'ai_checklist', 'ai_root', 'ai_chat_history'):
    if _k not in st.session_state: st.session_state[_k] = None


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:1.5rem;">
      <div style="width:32px;height:32px;
        background:linear-gradient(135deg,#1d4ed8,#7c3aed);
        border-radius:8px;display:flex;align-items:center;
        justify-content:center;font-size:1rem;">⚡</div>
      <div>
        <div style="font-size:0.78rem;font-weight:700;color:#e2e8f0;">DMaaS</div>
        <div style="font-size:0.65rem;color:#475569;letter-spacing:1px;
          text-transform:uppercase;">Prevalidation Engine</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<p style="font-size:0.65rem;text-transform:uppercase;letter-spacing:2px;'
                'color:#374151;font-weight:600;margin:0 0 6px;">Groq API Key</p>',
                unsafe_allow_html=True)
    sidebar_key = st.text_input("", type="password", placeholder="gsk_...",
                                value=st.session_state.get("_sidebar_key", ""),
                                label_visibility="collapsed")
    if sidebar_key: st.session_state["_sidebar_key"] = sidebar_key
    key_active = bool(_resolve_key())
    dot, col, txt = ("●", "#4ade80", "API key active") if key_active \
                 else ("○", "#f87171", "No key — AI disabled")
    st.markdown(f'<p style="font-size:0.72rem;color:{col};margin-top:4px;">'
                f'{dot} {txt}</p>', unsafe_allow_html=True)

    st.markdown('<hr style="border:none;border-top:1px solid #1e2a3a;margin:1rem 0;">',
                unsafe_allow_html=True)
    st.markdown("""
    <div style="background:#0a0f1a;border:1px solid #1e2a3a;border-radius:8px;padding:10px 12px;">
      <div style="color:#60a5fa;font-family:'JetBrains Mono',monospace;
          font-weight:600;font-size:0.7rem;margin-bottom:4px;">llama-3.3-70b-versatile</div>
      <div style="font-size:0.72rem;color:#475569;">Groq free tier · 14,400 req/day</div>
      <div style="font-size:0.72rem;color:#374151;margin-top:2px;">
        Fallback: llama-3.1-8b-instant</div>
    </div>
    <p style="font-size:0.72rem;color:#475569;margin-top:10px;">
      Get a free key at
      <a href="https://console.groq.com" target="_blank"
         style="color:#60a5fa;">console.groq.com</a>
    </p>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# HERO
# ─────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-wrap">
  <div>
    <div class="hero-brand">DMaaS · Data Migration as a Service</div>
    <div class="hero-title">Prevalidation Engine</div>
    <div class="hero-sub">Validate SAP migration data against field rules — before go-live.</div>
  </div>
  <div class="hero-badge">
    <div class="hero-dot"></div>Groq · Llama 3.3 70B
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# FILE UPLOAD
# ─────────────────────────────────────────────────────────────
st.markdown('<div class="sec-label">Input Files</div>', unsafe_allow_html=True)
c1, c2 = st.columns(2, gap="large")
with c1:
    st.markdown('<p style="font-size:0.75rem;color:#475569;margin-bottom:6px;font-weight:500;">'
                '📋  FIELD RULES FILE</p>', unsafe_allow_html=True)
    rules_file = st.file_uploader('Rules', type=['xlsx'], label_visibility='collapsed')
with c2:
    st.markdown('<p style="font-size:0.75rem;color:#475569;margin-bottom:6px;font-weight:500;">'
                '📊  MIGRATION DATA FILE</p>', unsafe_allow_html=True)
    product_file = st.file_uploader('Data', type=['xlsx'], label_visibility='collapsed')

if product_file is not None:
    fid = f"{product_file.name}_{product_file.size}"
    if st.session_state.get('_pfid') != fid:
        st.session_state['_pbytes'] = product_file.read()
        st.session_state['_pfid']   = fid

product_bytes_cached = st.session_state.get('_pbytes')
if product_bytes_cached:
    wb_tmp = load_workbook(BytesIO(product_bytes_cached), read_only=True)
    sheets_str = '  ·  '.join(wb_tmp.sheetnames); n = len(wb_tmp.sheetnames); wb_tmp.close()
    st.markdown(f'<p style="font-size:0.75rem;color:#4ade80;margin-top:8px;">'
                f'✓ {n} sheet(s) detected: {sheets_str}</p>', unsafe_allow_html=True)

st.markdown('<br>', unsafe_allow_html=True)

# ── Controls ──────────────────────────────────────────────────
ct1, ct2, ct3 = st.columns([2, 1, 2], gap="large")
with ct1:
    use_llm = st.toggle('🤖 Enable AI Insights', value=False,
                        help='Runs 5 AI features: Health Brief · Fix Guide · '
                             'Go/No-Go Checklist · Root Cause Analysis · Data Analyst Chat')
with ct2:
    run = st.button('▶  Run Validation', use_container_width=True)
with ct3:
    if use_llm and not key_active:
        st.markdown('<p style="font-size:0.75rem;color:#f87171;padding-top:10px;">'
                    '⚠️ Add a Groq API key in the sidebar to enable AI.</p>',
                    unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# VALIDATION RUN
# ─────────────────────────────────────────────────────────────
if run:
    if not rules_file or not product_bytes_cached:
        st.warning('Please upload both the Rules file and the Data file.'); st.stop()
    rules_bytes   = rules_file.read()
    product_bytes = product_bytes_cached
    base          = os.path.splitext(product_file.name)[0]
    log: list     = []; all_errors: list = []

    with st.spinner('Running validation …'):
        try:
            log.append('─── LOADING RULES ───────────────────────')
            all_rules = load_rules(rules_bytes, log)
            log.append(f'Rules: {list(all_rules.keys())}')
            wb_prod   = load_workbook(BytesIO(product_bytes), data_only=True)
            log.append(f'Sheets: {list(wb_prod.sheetnames)}')
            sheet_row_counts: dict = {}
            for sn in wb_prod.sheetnames:
                log.append(f'\n─── {sn} ─────────────────────────')
                errs, n_rows = validate_sheet(wb_prod[sn], sn, all_rules, log)
                all_errors.extend(errs); sheet_row_counts[sn] = n_rows
            log.append(f'\n─── TOTAL ERRORS: {len(all_errors)} ──────────────')
            DISPLAY  = ['Sheet Name', 'Row Number', 'Product Number', 'Field Name',
                        'Field Value', 'Error Type', 'Error Message']
            df_csv   = (pd.DataFrame(all_errors)[DISPLAY] if all_errors
                        else pd.DataFrame(columns=DISPLAY))
            run_ts       = datetime.now().strftime('%d-%m-%Y_%H-%M-%S')
            log_filename = f'{base}_{run_ts}.txt'
            row_risk     = compute_row_risk(all_errors)

            ai_brief_r = ai_fix_r = ai_chk_r = ai_root_r = None
            if use_llm and key_active:
                steps = [
                    ("Health Brief",        lambda: ai_health_brief(all_errors, sheet_row_counts, product_file.name)),
                    ("Fix Guide",           lambda: ai_fix_guide(all_errors)),
                    ("Readiness Checklist", lambda: ai_readiness_checklist(all_errors, sheet_row_counts, product_file.name)),
                    ("Root Cause Analysis", lambda: ai_root_cause(all_errors, sheet_row_counts)),
                ]
                prog = st.progress(0)
                results = []
                for i, (label, fn) in enumerate(steps):
                    prog.progress(int(i / len(steps) * 100), text=f"🤖 AI: {label} …")
                    results.append(fn())
                prog.progress(100, text="🤖 AI analysis complete ✅"); prog.empty()
                ai_brief_r, ai_fix_r, ai_chk_r, ai_root_r = results

            for k, v in {
                'all_errors': all_errors, 'log': log, 'base_name': base,
                'log_filename': log_filename, 'sheet_row_counts': sheet_row_counts,
                'row_risk': row_risk, 'ai_brief': ai_brief_r, 'ai_fix': ai_fix_r,
                'ai_checklist': ai_chk_r, 'ai_root': ai_root_r, 'ai_chat_history': [],
                'log_txt_bytes': '\n'.join(log).encode('utf-8'),
                'hl_bytes':  build_highlighted(product_bytes, all_errors, all_rules),
                'err_bytes': build_error_report(all_errors, product_file.name,
                                                sheet_row_counts, row_risk),
                'csv_bytes': df_csv.to_csv(index=False).encode('utf-8'),
            }.items():
                st.session_state[k] = v

        except Exception as ex:
            import traceback
            st.error(str(ex)); st.code(traceback.format_exc()); st.stop()


# ─────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────
if st.session_state.get('all_errors') is not None:
    all_errors       = st.session_state['all_errors']
    log              = st.session_state['log']
    base             = st.session_state['base_name']
    sheet_row_counts = st.session_state.get('sheet_row_counts') or {}
    row_risk         = st.session_state.get('row_risk') or {}

    st.markdown('<br>', unsafe_allow_html=True)
    st.markdown('<div class="sec-label">Validation Results</div>', unsafe_allow_html=True)

    # ── KPIs ─────────────────────────────────────────────────
    total_all = sum(sheet_row_counts.values())
    error_all = len(set((e['Sheet Name'], e['Row Number']) for e in all_errors))
    clean_all = total_all - error_all
    total_e   = len(all_errors)
    mand_e    = sum(1 for e in all_errors if e['Error Type'] == 'Mandatory Field Missing')
    fmt_e     = sum(1 for e in all_errors if e['Error Type'] in
                    ('Invalid Date Format', 'Invalid Time Format', 'Invalid Number', 'Length Exceeded'))
    clean_pct = round(clean_all / max(total_all, 1) * 100, 1)

    st.markdown(f"""
    <div class="kpi-grid">
      <div class="kpi red">
        <div class="kpi-num">{total_e:,}</div>
        <div class="kpi-lbl">Total Errors</div>
        <div class="kpi-sub">{error_all:,} rows affected</div>
      </div>
      <div class="kpi red">
        <div class="kpi-num">{mand_e:,}</div>
        <div class="kpi-lbl">Mandatory Missing</div>
        <div class="kpi-sub">Blocks migration</div>
      </div>
      <div class="kpi amber">
        <div class="kpi-num">{fmt_e:,}</div>
        <div class="kpi-lbl">Format &amp; Length</div>
        <div class="kpi-sub">Date · Number · Time · Length</div>
      </div>
      <div class="kpi green">
        <div class="kpi-num">{clean_pct}%</div>
        <div class="kpi-lbl">Clean Rows</div>
        <div class="kpi-sub">{clean_all:,} of {total_all:,}</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    if not all_errors:
        st.markdown('<div class="verdict-pass">✅ All records passed validation — '
                    'data is ready for migration.</div>', unsafe_allow_html=True)

    # ── Sheet Breakdown ───────────────────────────────────────
    st.markdown('<br>', unsafe_allow_html=True)
    st.markdown('<div class="sec-label">Sheet Breakdown</div>', unsafe_allow_html=True)
    sstats = defaultdict(lambda: defaultdict(int))
    for e in all_errors:
        sstats[e['Sheet Name']]['total']         += 1
        sstats[e['Sheet Name']][e['Error Type']] += 1
    rows_html = ''
    for sn in sheet_row_counts:
        tr  = sheet_row_counts.get(sn, 0)
        er  = len(set(e['Row Number'] for e in all_errors if e['Sheet Name'] == sn))
        sc  = sstats.get(sn, {})
        bdg = ('<span class="badge pass">✓ CLEAN</span>' if er == 0
               else f'<span class="badge fail">{er:,} error rows</span>')
        fmt = (sc.get('Invalid Date Format', 0) + sc.get('Invalid Time Format', 0)
               + sc.get('Invalid Number', 0))
        rows_html += (
            f'<tr><td><strong style="color:#e2e8f0;">{sn}</strong></td>'
            f'<td>{tr:,}</td><td>{(tr-er):,}</td><td>{bdg}</td>'
            f'<td style="color:#f87171;">{sc.get("Mandatory Field Missing",0) or "—"}</td>'
            f'<td style="color:#fbbf24;">{fmt or "—"}</td>'
            f'<td style="color:#a78bfa;">{sc.get("Length Exceeded",0) or "—"}</td></tr>'
        )
    st.markdown(f"""
    <table class="sheet-table">
      <thead><tr>
        <th>Sheet</th><th>Total Rows</th><th>Clean Rows</th><th>Status</th>
        <th>Mandatory</th><th>Format Errors</th><th>Length</th>
      </tr></thead>
      <tbody>{rows_html}</tbody>
    </table>
    """, unsafe_allow_html=True)

    # ── High-Risk Rows ────────────────────────────────────────
    if row_risk:
        high = sorted([(k, v) for k, v in row_risk.items() if v >= 60],
                      key=lambda x: -x[1])[:8]
        if high:
            st.markdown('<br>', unsafe_allow_html=True)
            st.markdown('<div class="sec-label">High-Risk Rows</div>', unsafe_allow_html=True)
            st.markdown('<p style="font-size:0.75rem;color:#475569;margin-bottom:10px;">'
                        'Rows with highest combined error severity. Prioritise these.</p>',
                        unsafe_allow_html=True)
            for (sn, rn), score in high:
                cls = 'high' if score >= 80 else 'med'
                st.markdown(f"""
                <div class="risk-row">
                  <div style="font-size:0.75rem;color:#94a3b8;width:200px;flex-shrink:0;">
                    <strong style="color:#e2e8f0;">{sn}</strong> · Row {rn}
                  </div>
                  <div class="risk-bar-wrap">
                    <div class="risk-bar {cls}" style="width:{score}%;"></div>
                  </div>
                  <div class="risk-meta">{score}/100</div>
                </div>
                """, unsafe_allow_html=True)

    # ── AI Insights ───────────────────────────────────────────
    ai_brief_val = st.session_state.get('ai_brief')
    ai_fix_val   = st.session_state.get('ai_fix')
    ai_chk_val   = st.session_state.get('ai_checklist')
    ai_root_val  = st.session_state.get('ai_root')
    ai_hist      = st.session_state.get('ai_chat_history') or []
    has_ai = any([ai_brief_val, ai_fix_val, ai_chk_val, ai_root_val])

    if has_ai or use_llm:
        st.markdown('<br>', unsafe_allow_html=True)
        st.markdown('<div class="sec-label">AI Insights · Groq Llama 3.3 70B</div>',
                    unsafe_allow_html=True)
        tabs = st.tabs(['📋 Health Brief', '🔧 Fix Guide',
                        '✅ Go / No-Go', '🔍 Root Cause', '💬 Ask the Data'])

        # T1: Health Brief ─────────────────────────────────────
        with tabs[0]:
            if ai_brief_val:
                lines = [l.strip() for l in ai_brief_val.strip().split('\n') if l.strip()]
                body  = '<br>'.join(
                    f"<strong style='color:#e2e8f0;'>{l}</strong>" if i == 0 else l
                    for i, l in enumerate(lines))
                st.markdown(f'<div class="ai-card"><div class="ai-card-title">'
                            f'📋 Migration Health Brief</div>'
                            f'<div class="ai-card-body">{body}</div></div>',
                            unsafe_allow_html=True)
            else:
                st.caption('Enable AI Insights and run validation to generate the health brief.')

        # T2: Fix Guide ────────────────────────────────────────
        with tabs[1]:
            if ai_fix_val and isinstance(ai_fix_val, dict):
                st.markdown('<p style="font-size:0.78rem;color:#475569;margin-bottom:1rem;">'
                            'Root cause and fix guidance for the top error patterns.</p>',
                            unsafe_allow_html=True)
                for (etype, field), info in ai_fix_val.items():
                    bc = ('mand' if 'Mandatory' in etype else
                          'date' if 'Date' in etype or 'Time' in etype else
                          'num'  if 'Number' in etype else 'len')
                    bl = (etype.replace('Mandatory Field Missing', 'MANDATORY')
                               .replace('Invalid Date Format', 'DATE')
                               .replace('Invalid Time Format', 'TIME')
                               .replace('Invalid Number', 'NUMBER')
                               .replace('Length Exceeded', 'LENGTH'))
                    why    = info.get('why',    '') if isinstance(info, dict) else str(info)
                    fix    = info.get('fix',    '') if isinstance(info, dict) else ''
                    verify = info.get('verify', '') if isinstance(info, dict) else ''
                    st.markdown(f"""
                    <div class="fix-card">
                      <div class="fix-header">
                        <span class="fix-badge {bc}">{bl}</span>
                        <span class="fix-field">{field}</span>
                      </div>
                      <div class="fix-row"><strong>Why:</strong> {why}</div>
                      <div class="fix-row"><strong>Fix:</strong> {fix}</div>
                      <div class="fix-row"><strong>Verify:</strong> {verify}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.caption('No fix guide generated.')

        # T3: Go / No-Go Checklist ─────────────────────────────
        with tabs[2]:
            if ai_chk_val and isinstance(ai_chk_val, list):
                passes = sum(1 for x in ai_chk_val if x.get('status') == 'PASS')
                fails  = sum(1 for x in ai_chk_val if x.get('status') == 'FAIL')
                warns  = sum(1 for x in ai_chk_val if x.get('status') == 'WARN')
                if fails == 0 and warns == 0:
                    vcls, vtxt = 'verdict-pass', '✅  GO — Data quality meets migration threshold.'
                elif fails == 0:
                    vcls, vtxt = 'verdict-warn', '⚠️  CONDITIONAL GO — Resolve warnings before scheduling go-live.'
                else:
                    vcls = 'verdict-fail'
                    vtxt = f'❌  NO-GO — {fails} critical issue(s) must be resolved before migration.'
                st.markdown(f'<div class="{vcls}">{vtxt}</div>', unsafe_allow_html=True)
                st.markdown('<br>', unsafe_allow_html=True)
                for item in ai_chk_val:
                    s   = item.get('status', 'WARN')
                    ico = {'PASS': '✅', 'FAIL': '❌', 'WARN': '⚠️'}.get(s, '⚠️')
                    st.markdown(f"""
                    <div class="chk-item {s.lower()}">
                      <div class="chk-status">{ico}</div>
                      <div>
                        <div class="chk-title">{item.get('item','')}</div>
                        <div class="chk-reason">{item.get('reason','')}</div>
                      </div>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown(f'<p style="font-size:0.72rem;color:#374151;margin-top:0.8rem;">'
                            f'{passes} PASS · {warns} WARN · {fails} FAIL</p>',
                            unsafe_allow_html=True)
            else:
                st.caption('No checklist generated.')

        # T4: Root Cause Analysis ──────────────────────────────
        with tabs[3]:
            if ai_root_val:
                body = ai_root_val.strip().replace('\n\n', '<br><br>').replace('\n', '<br>')
                st.markdown(f'<div class="ai-card"><div class="ai-card-title">'
                            f'🔍 Root Cause Analysis</div>'
                            f'<div class="ai-card-body">{body}</div></div>',
                            unsafe_allow_html=True)
            else:
                st.caption('No root cause analysis generated.')

        # T5: Data Analyst Chat ────────────────────────────────
        with tabs[4]:
            st.markdown(
                '<p style="font-size:0.78rem;color:#475569;margin-bottom:1rem;">'
                'Ask anything about your results. Try: '
                '<em>"Which sheet has the most mandatory errors?", '
                '"What % of rows are clean?", '
                '"Why are there so many date errors?"</em></p>',
                unsafe_allow_html=True)
            for msg in ai_hist:
                with st.chat_message(msg['role']): st.write(msg['content'])
            user_q = st.chat_input('Ask about your migration data …')
            if user_q:
                if not key_active:
                    st.warning('Add a Groq API key in the sidebar to use the chat.')
                else:
                    ai_hist.append({'role': 'user', 'content': user_q})
                    with st.chat_message('user'): st.write(user_q)
                    with st.chat_message('assistant'):
                        with st.spinner('Analysing …'):
                            ans = ai_chat(user_q, all_errors, sheet_row_counts,
                                          st.session_state.get('base_name', 'file'))
                        st.write(ans)
                    ai_hist.append({'role': 'assistant', 'content': ans})
                    st.session_state['ai_chat_history'] = ai_hist

    # ── Validation Log ────────────────────────────────────────
    st.markdown('<br>', unsafe_allow_html=True)
    st.markdown('<div class="sec-label">Validation Log</div>', unsafe_allow_html=True)
    log_fname = st.session_state.get('log_filename', 'validation_log.txt')
    with st.expander(f'📋 {log_fname}', expanded=False):
        st.markdown(f'<div class="log-box">{"<br>".join(log)}</div>', unsafe_allow_html=True)

    # ── Downloads ─────────────────────────────────────────────
    st.markdown('<br>', unsafe_allow_html=True)
    st.markdown('<div class="sec-label">Export Results</div>', unsafe_allow_html=True)
    d1, d2, d3, d4 = st.columns(4, gap="medium")
    with d1:
        st.download_button('⬇  Highlighted Data',
            data=st.session_state['hl_bytes'],
            file_name=f'Highlighted_{base}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True)
    with d2:
        st.download_button('⬇  Error Report (.xlsx)',
            data=st.session_state['err_bytes'],
            file_name=f'ErrorReport_{base}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True)
    with d3:
        st.download_button('⬇  Error Log (.csv)',
            data=st.session_state['csv_bytes'],
            file_name=f'ErrorLog_{base}.csv',
            mime='text/csv',
            use_container_width=True)
    with d4:
        st.download_button('⬇  Validation Log (.txt)',
            data=st.session_state.get('log_txt_bytes', b''),
            file_name=log_fname,
            mime='text/plain',
            use_container_width=True)